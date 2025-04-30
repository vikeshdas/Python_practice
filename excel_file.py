"""
openpyxl is a Python library specifically for reading and writing Excel .xlsx files. It can read, write, modify, and style Excel files directly

openpyxl work only with .xlsx file 
"""

from openpyxl import load_workbook

workbook = load_workbook('generated_file.xlsx')

# there can be multiple subsheet in excell file so reading current open sheet
sheet = workbook.active


for row in sheet.iter_rows(values_only=True):
    print(row)

